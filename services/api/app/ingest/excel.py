"""Read ДЗК / ДРРП workbooks and emit canonical row dicts.

Uses ``openpyxl`` in read-only, data-only mode. The real files contain ~21k and
~20k rows respectively, which openpyxl handles in <2 s on a laptop.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

import openpyxl

from app.ingest.normalize import (
    excel_serial_to_date,
    ga_to_m2,
    normalize_address,
    normalize_cadastral,
    normalize_koatuu,
    normalize_object_type,
    normalize_tax_id,
    parse_float,
    split_intended_use,
    valuation_to_kop,
)

_ZEM_HEADERS = {
    "Кадастровий номер",
    "koatuu",
    "Форма власності",
    "Цільове призначення",
    "Місцерозташування",
    "Вид с/г угідь",
    "Площа, га",
    "Усереднена нормативно грошова оцінка",
    "ЄДРПОУ землекористувача",
    "Землекористувач",
    "Частка володіння",
    "Дата державної реєстрації права власності",
    "Номер запису про право власності",
    "Орган, що здійснив державну реєстрацію права власності",
    "Тип",
    "Підтип",
}

_NER_HEADERS = {
    "Податковий номер ПП",
    "Назва платника",
    "Тип об'єкта",
    "Адреса об'єкта",
    "Дата держ. реєстр. права власн",
    "Дата держ. реєстр. прип. права власн",
    "Загальна площа",
    "Розмір частки у праві спільної власності",
}


def _open_sheet(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return wb, wb[wb.sheetnames[0]]


def _read_headers(ws) -> list[str]:
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(cell).strip() if cell is not None else "" for cell in first_row]


def _iter_records(path: Path, required: set[str]) -> Iterator[dict[str, Any]]:
    wb, ws = _open_sheet(path)
    try:
        headers = _read_headers(ws)
        missing = required - set(headers)
        if missing:
            raise ValueError(f"Missing expected columns in {path.name}: {sorted(missing)}")
        indexes = {h: i for i, h in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            yield {h: row[indexes[h]] if indexes[h] < len(row) else None for h in headers}
    finally:
        wb.close()


def read_zem_workbook(path: str | Path) -> Iterator[dict[str, Any]]:
    """Yield normalized ДЗК rows keyed by canonical snake_case field names."""
    p = Path(path)
    for raw in _iter_records(p, _ZEM_HEADERS):
        cad = normalize_cadastral(raw.get("Кадастровий номер"))
        if not cad:
            continue
        code, label = split_intended_use(raw.get("Цільове призначення"))
        yield {
            "cadastral_no": cad,
            "koatuu": normalize_koatuu(raw.get("koatuu")),
            "ownership_form": (raw.get("Форма власності") or None) and str(raw["Форма власності"]).strip().lower() or None,
            "intended_use_code": code,
            "intended_use_label": label,
            "location_admin": (raw.get("Місцерозташування") or None) and str(raw["Місцерозташування"]).strip() or None,
            "agri_use_kind": (raw.get("Вид с/г угідь") or None) and str(raw["Вид с/г угідь"]).strip() or None,
            "area_m2": ga_to_m2(raw.get("Площа, га")),
            "valuation_kop": valuation_to_kop(raw.get("Усереднена нормативно грошова оцінка")),
            "owner_tax_id": normalize_tax_id(raw.get("ЄДРПОУ землекористувача")),
            "owner_name_raw": (raw.get("Землекористувач") or None) and str(raw["Землекористувач"]).strip() or None,
            "share": parse_float(raw.get("Частка володіння"), default=1.0),
            "registered_at": excel_serial_to_date(raw.get("Дата державної реєстрації права власності")),
            "record_no": (raw.get("Номер запису про право власності") or None)
            and str(raw["Номер запису про право власності"]).strip()
            or None,
            "registrar": (raw.get("Орган, що здійснив державну реєстрацію права власності") or None)
            and str(raw["Орган, що здійснив державну реєстрацію права власності"]).strip()
            or None,
            "record_kind": (raw.get("Тип") or None) and str(raw["Тип"]).strip() or None,
            "record_subkind": (raw.get("Підтип") or None) and str(raw["Підтип"]).strip() or None,
        }


def read_ner_workbook(path: str | Path) -> Iterator[dict[str, Any]]:
    """Yield normalized ДРРП rows keyed by canonical snake_case field names."""
    p = Path(path)
    # ДРРП header "Вид спіль ної власності" has a stray space in the real file; we don't require it.
    for raw in _iter_records(p, _NER_HEADERS):
        tax_id = normalize_tax_id(raw.get("Податковий номер ПП"))
        if not tax_id:
            continue
        object_raw_val = raw.get("Тип об'єкта")
        object_raw = str(object_raw_val).strip() if object_raw_val else None
        address_raw_val = raw.get("Адреса об'єкта")
        address_raw = str(address_raw_val).strip() if address_raw_val else None
        joint_val = raw.get("Вид спіль ної власності") or raw.get("Вид спільної власності")
        yield {
            "owner_tax_id": tax_id,
            "owner_name_raw": (raw.get("Назва платника") or None) and str(raw["Назва платника"]).strip() or None,
            "object_type_raw": object_raw,
            "object_type_norm": normalize_object_type(object_raw),
            "address_raw": address_raw,
            "address_norm": normalize_address(address_raw),
            "area_m2": parse_float(raw.get("Загальна площа"), default=0.0),
            "registered_at": excel_serial_to_date(raw.get("Дата держ. реєстр. права власн")),
            "terminated_at": excel_serial_to_date(raw.get("Дата держ. реєстр. прип. права власн")),
            "joint_ownership_kind": (str(joint_val).strip().lower() if joint_val else None) or None,
            "share": parse_float(raw.get("Розмір частки у праві спільної власності"), default=1.0),
        }
