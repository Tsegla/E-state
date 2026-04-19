"""Yearly subscription quote engine for OTG land audits."""

from __future__ import annotations

import csv
import math
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.schemas import SubscriptionQuoteDTO
from app.db.models import DatasetRow, LandParcelRow
from app.matcher.config import default_config as default_matcher_config
from app.pricing.config import PricingConfig, default_config

OWNER_EMPTY_MARKERS = {"", "nan", "null", "невідомо", "невiдомо", "none"}
DATASET_ROW_LIMIT = 200_000

_MATCHER_CFG = default_matcher_config()

HEADER_ALIASES = {
    "cadastral_no": {
        "Кадастровий номер",
        "кадастровий номер",
        "cadastral_no",
        "cadastral number",
    },
    "owner_id": {
        "ЄДРПОУ землекористувача",
        "Податковий номер ПП",
        "owner_tax_id",
        "owner_id",
        "tax_id",
    },
    "owner_name": {
        "Землекористувач",
        "Назва платника",
        "owner_name",
        "owner",
    },
    "area_ha": {"Площа, га", "area_ha", "area (ha)", "площа га"},
    "ngo_uah_per_ha": {
        "Усереднена нормативно грошова оцінка",
        "НГО",
        "ngo",
        "ngo_uah_per_ha",
    },
    "intended_use_code": {"intended_use_code", "Код цільового призначення", "код цільового"},
    "intended_use_label": {"Цільове призначення", "intended_use", "цільове призначення"},
}


@dataclass(frozen=True, slots=True)
class NormalizedParcel:
    cadastral_no: str | None
    owner_id: str
    area_ha: float
    ngo_uah_per_ha: float
    intended_use_code: str | None
    intended_use_label: str | None


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _coerce_text(value).replace(" ", "")
    if not text:
        return 0.0
    normalized = text.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def _normalize_owner(owner_id: Any, owner_name: Any, *, cfg: PricingConfig) -> str:
    identifier = _coerce_text(owner_id) or _coerce_text(owner_name)
    normalized = identifier.lower()
    if normalized in OWNER_EMPTY_MARKERS:
        return cfg.missing_owner_sentinel
    return identifier if identifier else cfg.missing_owner_sentinel


def _normalize_use(raw_use: str | None) -> tuple[str | None, str | None]:
    if not raw_use:
        return (None, None)
    text = raw_use.strip()
    if not text:
        return (None, None)
    code = None
    label = text
    if " " in text:
        first, rest = text.split(" ", 1)
        if len(first) == 5 and first[2] == ".":
            code = first
            label = rest.strip() or text
    return code, label


def _parse_row(raw: dict[str, Any], *, cfg: PricingConfig) -> NormalizedParcel:
    area_ha = max(_coerce_number(raw.get("area_ha")), 0.0)
    ngo = max(_coerce_number(raw.get("ngo_uah_per_ha")), 0.0)
    owner = _normalize_owner(raw.get("owner_id"), raw.get("owner_name"), cfg=cfg)

    raw_code = _coerce_text(raw.get("intended_use_code")) or None
    raw_label = _coerce_text(raw.get("intended_use_label")) or None
    if raw_code is None and raw_label:
        parsed_code, parsed_label = _normalize_use(raw_label)
        raw_code = parsed_code
        raw_label = parsed_label

    cadastral = _coerce_text(raw.get("cadastral_no")) or None
    return NormalizedParcel(
        cadastral_no=cadastral,
        owner_id=owner,
        area_ha=area_ha,
        ngo_uah_per_ha=ngo,
        intended_use_code=raw_code,
        intended_use_label=raw_label,
    )


def _dedupe(rows: list[NormalizedParcel]) -> list[NormalizedParcel]:
    has_cadastral = any(row.cadastral_no for row in rows)
    seen: set[tuple[Any, ...]] = set()
    unique: list[NormalizedParcel] = []
    for row in rows:
        key: tuple[Any, ...]
        if has_cadastral and row.cadastral_no:
            key = (row.cadastral_no or "",)
        else:
            key = (
                row.owner_id,
                round(row.area_ha, 4),
                round(row.ngo_uah_per_ha, 2),
                row.intended_use_code or row.intended_use_label or "",
            )
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _classify_purpose(intended_use_code: str | None) -> str:
    if not intended_use_code:
        return "other"
    if intended_use_code in _MATCHER_CFG.agri_use_codes:
        return "agri"
    if intended_use_code in _MATCHER_CFG.residential_use_codes:
        return "residential"
    if intended_use_code in _MATCHER_CFG.commercial_use_codes:
        return "commercial"
    if intended_use_code in _MATCHER_CFG.industrial_use_codes:
        return "industrial"
    return "other"


def _concentration_multiplier(rows: list[NormalizedParcel], *, cfg: PricingConfig) -> tuple[float, float]:
    owner_area: dict[str, float] = {}
    total_area = 0.0
    for row in rows:
        total_area += row.area_ha
        if row.owner_id == cfg.missing_owner_sentinel:
            continue
        owner_area[row.owner_id] = owner_area.get(row.owner_id, 0.0) + row.area_ha

    if total_area <= 0 or not owner_area:
        return (1.0, 0.0)

    areas = sorted(owner_area.values(), reverse=True)
    top_n = max(1, math.ceil(len(areas) * 0.1))
    top_share = max(min(sum(areas[:top_n]) / total_area, 1.0), 0.0)
    for threshold, multiplier in cfg.concentration_thresholds:
        if top_share >= threshold:
            return (multiplier, top_share)
    raise AssertionError("concentration_thresholds must include a 0.0 fallback threshold")


def _tier_for_multiplier(multiplier: float) -> str:
    if multiplier >= 1.6:
        return "premium"
    if multiplier >= 1.3:
        return "mid"
    return "base"


def compute_quote(
    raw_rows: list[dict[str, Any]],
    *,
    cfg: PricingConfig | None = None,
    dataset_id: UUID | None = None,
) -> SubscriptionQuoteDTO:
    pricing_cfg = cfg or default_config()
    validated = [_parse_row(row, cfg=pricing_cfg) for row in raw_rows]
    deduped = _dedupe(validated)

    concentration_multiplier, top_share = _concentration_multiplier(deduped, cfg=pricing_cfg)
    revenue_by_purpose: dict[str, float] = {}
    owners = {row.owner_id for row in deduped if row.owner_id != pricing_cfg.missing_owner_sentinel}
    total_area = 0.0
    projected_revenue = 0.0

    for row in deduped:
        purpose = _classify_purpose(row.intended_use_code)
        purpose_multiplier = pricing_cfg.purpose_multipliers.get(purpose, 1.0)
        ngo = row.ngo_uah_per_ha if row.ngo_uah_per_ha > 0 else pricing_cfg.fallback_ngo_uah_per_ha
        updated_ngo = ngo * pricing_cfg.indexation_year_coeff
        row_revenue = row.area_ha * updated_ngo * pricing_cfg.base_land_tax_rate * purpose_multiplier
        projected_revenue += row_revenue
        revenue_by_purpose[purpose] = revenue_by_purpose.get(purpose, 0.0) + row_revenue
        total_area += row.area_ha

    projected_revenue *= concentration_multiplier
    raw_subscription = projected_revenue * pricing_cfg.subscription_share
    yearly_price = max(
        pricing_cfg.subscription_floor_uah,
        min(pricing_cfg.subscription_cap_uah, raw_subscription),
    )

    return SubscriptionQuoteDTO(
        yearly_price_uah=round(yearly_price, 2),
        projected_recoverable_revenue_uah=round(projected_revenue, 2),
        concentration_multiplier=round(concentration_multiplier, 3),
        top10_percent_area_share=round(top_share, 4),
        tier=_tier_for_multiplier(concentration_multiplier),
        total_parcels=len(deduped),
        total_owners=len(owners),
        total_area_ha=round(total_area, 4),
        revenue_by_purpose={key: round(value, 2) for key, value in revenue_by_purpose.items()},
        inputs={
            "base_land_tax_rate": pricing_cfg.base_land_tax_rate,
            "indexation_year_coeff": pricing_cfg.indexation_year_coeff,
            "subscription_share": pricing_cfg.subscription_share,
            "subscription_floor_uah": pricing_cfg.subscription_floor_uah,
            "subscription_cap_uah": pricing_cfg.subscription_cap_uah,
            "fallback_ngo_uah_per_ha": pricing_cfg.fallback_ngo_uah_per_ha,
        },
        caveats=[
            "Це розрахунок орієнтовної річної ціни підписки, а не юридичне податкове донарахування.",
            "Якщо НГО відсутня або некоректна, використовується регіональна ставка за замовчуванням.",
        ],
        generated_at=datetime.now(tz=timezone.utc),
        dataset_id=dataset_id,
    )


def _normalize_header(header: str) -> str:
    return " ".join(header.replace("_", " ").strip().lower().split())


def _header_mapping(headers: list[str]) -> dict[str, str]:
    normalized_to_raw = {_normalize_header(header): header for header in headers}
    mapping: dict[str, str] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = normalized_to_raw.get(_normalize_header(alias))
            if key:
                mapping[canonical] = key
                break
    return mapping


def _materialize_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not records:
        return []
    mapping = _header_mapping(list(records[0].keys()))
    output: list[dict[str, Any]] = []
    for record in records:
        output.append(
            {
                "cadastral_no": record.get(mapping.get("cadastral_no", ""), None),
                "owner_id": record.get(mapping.get("owner_id", ""), None),
                "owner_name": record.get(mapping.get("owner_name", ""), None),
                "area_ha": record.get(mapping.get("area_ha", ""), None),
                "ngo_uah_per_ha": record.get(mapping.get("ngo_uah_per_ha", ""), None),
                "intended_use_code": record.get(mapping.get("intended_use_code", ""), None),
                "intended_use_label": record.get(mapping.get("intended_use_label", ""), None),
            }
        )
    return output


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader if row]


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            if row is None:
                continue
            if all(cell is None or _coerce_text(cell) == "" for cell in row):
                continue
            record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            records.append(record)
        return records
    finally:
        workbook.close()


def quote_from_file(path: str | Path, *, cfg: PricingConfig | None = None) -> SubscriptionQuoteDTO:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        records = _read_csv(file_path)
    elif suffix in {".xlsx", ".xls"}:
        records = _read_xlsx(file_path)
    else:
        raise ValueError("Unsupported format. Expected .xlsx, .xls, or .csv")
    return compute_quote(_materialize_rows(records), cfg=cfg)


def quote_from_dataset(
    session: Session, *, dataset_id: UUID, cfg: PricingConfig | None = None
) -> SubscriptionQuoteDTO:
    dataset = session.get(DatasetRow, dataset_id)
    if dataset is None:
        raise ValueError("dataset not found")

    rows = (
        session.execute(
            select(LandParcelRow)
            .where(LandParcelRow.dataset_id == dataset_id)
            .limit(DATASET_ROW_LIMIT)
        )
        .scalars()
        .all()
    )
    raw_rows: list[dict[str, Any]] = []
    for row in rows:
        raw_rows.append(
            {
                "cadastral_no": row.cadastral_no,
                "owner_id": row.owner_tax_id,
                "owner_name": row.owner_name_raw,
                "area_ha": (row.area_m2 or 0.0) / 10_000.0,
                "ngo_uah_per_ha": (row.valuation_kop or 0) / 100.0 if row.valuation_kop else 0.0,
                "intended_use_code": row.intended_use_code,
                "intended_use_label": row.intended_use_label,
            }
        )
    return compute_quote(raw_rows, cfg=cfg, dataset_id=dataset_id)
